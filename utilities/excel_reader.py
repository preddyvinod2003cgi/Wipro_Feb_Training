import openpyxl


def get_booking_data(file):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    data = []
    headers = []

    # Read headers from first row
    for col in range(1, sheet.max_column + 1):
        headers.append(sheet.cell(row=1, column=col).value)

    # Read data rows
    for row in range(2, sheet.max_row + 1):

        row_data = {}

        for col in range(1, sheet.max_column + 1):

            key = headers[col - 1]
            value = sheet.cell(row=row, column=col).value

            row_data[key] = value

        data.append(row_data)

    return data